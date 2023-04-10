from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import win32com.client as win32
import os
from datetime import datetime
import pyexcel

from TV7Parser import TV7Parser
from SPTParser import SPTParser
from HeadData import *



class VKTParser:
    def __init__(self, data_list, curr_dir, save_dir):
        self.report = ''
        self.my_parsing_files = []
        self.my_dir = curr_dir
        self.save_dir = save_dir

        for file in data_list:
            if '.txt' in file:
                with open(file, 'r') as curr:
                    for line in curr:
                        if "Заводской номер" in line:
                            factor_num = line.split("Заводской номер")[1].split("ВВОД")[0].replace(' ', '')
                            input_type = line.split("ВВОД")[1].split("СХЕМА ПОДКЛЮЧЕНИЯ")[0].replace(' ', '')
                            if get_head_data(self.my_dir, factor_num, input_type)['type'] != None:
                                if 'ВКТ' in get_head_data(self.my_dir, factor_num, input_type)['type']:
                                    self.my_parsing_files.append([file, open(file, 'r')])
                                else:
                                    continue
                            else:
                                print('Can not to expect type in: ',  str(file) + '\n')
                        else:
                            continue

            elif 'xls' in file and 'xlsx' not in file:
                pyexcel.save_book_as(file_name=file,
                            dest_file_name=file + 'x')
                file += 'x'

            elif 'xlsx' in file:
                head_data = {'factory_num': '', 'complex_num': '', 'consumer': '', 'order': '', 'adress': '', 'cold_temp': '5,0', 'save_folder': '', 'type': ''}
                ws = load_workbook(file).active
                rep_type = '1'
                for row in ws.iter_rows(max_row=14):
                    if row[0].value == None:
                        continue
                    
                    # If visual type is as SPT
                    if '_1' in str(row[0].value):
                        head_data = get_head_data(self.my_dir, str(row[0].value), '1')
                        continue
                    if '_2' in str(row[0].value):
                        head_data = get_head_data(self.my_dir, str(row[0].value), '2')
                        continue

                    if 'ВВОД' in str(row[0].value):
                        if row[0].value.split('ВВОД')[1].split('СХЕМА')[0].replace(' ', '') == '2':
                            rep_type = '2'
                    
                    if 'Заводской номер' in str(row[0].value):
                        head_data = get_head_data(self.my_dir, str(row[0].value).split('Заводской номер')[1].split('ВВОД')[0].replace(' ', ''), rep_type)
                        print(str(row[0].value).split('Заводской номер')[1].split(' ')[0])
                if head_data['type'] != None:
                    if 'ВКТ' in str(head_data['type']):
                        self.my_parsing_files.append([file, load_workbook(file).active])
                    else:
                        continue
                else:
                    print('Can not to expect type in: ',  str(file))
            else:
                print('VKT wrong file!')
                continue 
            

    def data_index(self, line):
        nums_of_data = {'Дата': -1, 't1': -1, 't2': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Qо': -1, 'BНP': -1, 'BOC': -1}
        ind = 0
        for cell in line:
            for key in nums_of_data:
                if key in cell:
                    nums_of_data[key] = ind
            ind += 1
        return nums_of_data


    def build_xls_by_txt(self, file, date_from='', date_to=''):
        report_ = ''
        template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx', data_only=False)  # Template xlsx file
        if file[0].split('/')[len(file[0].split('/')) - 1]:
            template.title = file[0].split('/')[len(file[0].split('/')) - 1]
        ws = template.active
        ws['A1'] = str(ws['A1'].value).replace('май', get_month(datetime.now().strftime("%d-%m-%Y")))
        ws['B3'] = date_from
        ws['C3'] = date_to
        ws['B4'] = datetime.now().strftime("%d-%m-%Y")
        
        index = 1
        row_number = 18 
        head_data = {}
        input_type = ''
        with open(file[0], 'r') as curr:
            indexes = {}
            dm_sum = 0
            m1_sum = 0
            m2_sum = 0 
            dv_sum = 0
            v1_sum = 0
            v2_sum = 0 
            t1_avg = 0
            t2_avg = 0
            q_sum = 0
            vnr = 0
            vos = 0
            for line in curr:
                line_list = line.split('|')
                if "Заводской номер" in line:
                    factor_num = line.split("Заводской номер")[1].split("ВВОД")[0].replace(' ', '')
                    input_type = line.split("ВВОД")[1].split("СХЕМА ПОДКЛЮЧЕНИЯ")[0].replace(' ', '')
                    head_data = get_head_data(self.my_dir, factor_num, input_type)
                    ws['B5'] = input_type
                if '  Дaтa  ' == line_list[0]:
                    indexes = self.data_index(line_list)
                if index < 17:
                    index += 1
                    continue
                if len(line_list) <= 4:
                    break
                
                tmp_date = line_list[0].split('/')
                curr_date = datetime.strptime(tmp_date[0] + '-' + tmp_date[1] + '-' + '20' + tmp_date[2], "%d-%m-%Y").date()
                if date_from == '':
                    date_from = curr_date
                date_to = curr_date
                num = lambda st: float(line_list[indexes[st]].replace(' ', ''))
                # Inserting new row and styling it
                ws.insert_rows(row_number)
                for i in range(1, 14):
                    thin = Side(border_style="thin", color="000000")
                    ws.cell(row_number, i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(row_number, i).alignment = Alignment(horizontal="center", vertical="center")
                # A main table
                ws['A' + str(row_number)] = str(curr_date.strftime("%d-%m-%Y"))      #Date
                if indexes['t1'] == -1: ws['B' + str(row_number)] = ' - ' 
                else: 
                    ws['B' + str(row_number)] = str(num('t1')).replace('.', ',')
                    t1_avg += num('t1')   #T1
                if indexes['t2'] == -1: ws['C' + str(row_number)] = ' - ' 
                else: 
                    ws['C' + str(row_number)] = str(num('t2')).replace('.', ',')
                    t2_avg += num('t2')   #T2
                if indexes['V1'] == -1: ws['D' + str(row_number)] = ' - ' 
                else: 
                    ws['D' + str(row_number)] = str(num('V1')).replace('.', ',')
                    v1_sum += float(num('V1'))
                    ws['D' + str(row_number+1)] = str(round(v1_sum, 2)).replace('.', ',')
                if indexes['M1'] == -1: ws['E' + str(row_number)] = ' - ' 
                else: 
                    ws['E' + str(row_number)] = str(num('M1')).replace('.', ',')
                    m1_sum += num('M1')
                    ws['E' + str(row_number+1)] = str(round(m1_sum, 2)).replace('.', ',')
                if indexes['V2'] == -1: ws['F' + str(row_number)] = ' - ' 
                else: 
                    ws['F' + str(row_number)] = str(num('V2')).replace('.', ',')
                    v2_sum += num('V2')
                    ws['F' + str(row_number + 1)] = str(round(v2_sum, 2)).replace('.', ',')
                if indexes['M2'] == -1: ws['G' + str(row_number)] = ' - ' 
                else: 
                    ws['G' + str(row_number)] = str(num('M2')).replace('.', ',')
                    m2_sum += num('M2')
                    ws['G' + str(row_number + 1)] = str(round(m2_sum, 2)).replace('.', ',')
                if indexes['M2'] == -1 or indexes['V2'] == -1:
                    ws['H' + str(row_number)] = ws['D' + str(row_number)].value
                    ws['I' + str(row_number)] = ws['E' + str(row_number)].value
                    dv_sum = v1_sum
                    dm_sum = m1_sum
                    ws['H' + str(row_number + 1)] = dv_sum
                    ws['I' + str(row_number + 1)] = dm_sum
                else: 
                    ws['H' + str(row_number)] = str(round(num('V1') - num('V2')), 2).replace('.', ',')
                    dv_sum += round(abs(num('V2') - num('V1')), 2)
                    ws['I' + str(row_number)] = str(round(num('M1') - num('M2')), 2).replace('.', ',')
                    dm_sum += round(abs(num('M2') - num('M1')), 2)
                    ws['H' + str(row_number + 1)] = dv_sum
                    ws['I' + str(row_number + 1)] = dm_sum
                if indexes['Qо'] == -1: ws['J' + str(row_number)] = ' - ' 
                else: 
                    ws['J' + str(row_number)] = str(num('Qо')).replace('.', ',')
                    q_sum += num('Qо')
                    ws['J' + str(row_number + 1)] = str(round(q_sum, 3)).replace('.', ',')
                if indexes['BНP'] == -1: ws['K' + str(row_number)] = ' - ' 
                else: 
                    ws['K' + str(row_number)] = str(num('BНP')).replace('.', ',')
                    vnr += num('BНP')
                    ws['K' + str(row_number + 1)] = str(round(vnr, 2)).replace('.', ',')
                if indexes['BOC'] == -1: ws['L' + str(row_number)] = ' - ' 
                else: 
                    ws['L' + str(row_number)] = str(num('BOC')).replace('.', ',')
                    vos += num('BOC')
                    ws['L' + str(row_number + 1)] = str(round(vos, 2)).replace('.', ',')
                ws['M' + str(row_number)] = ' - ' #str(float(line_list[8].replace(' ', '')))
                row_number += 1
                index += 1
            
            if indexes['t1'] != -1:
                ws['B' + str(row_number+ 1)] = str(round(t1_avg / (index-18), 2)).replace('.', ',')
            if indexes['t2'] != -1:
                ws['C' + str(row_number+ 1)] = str(round(t2_avg / (index-18), 2)).replace('.', ',')
            
            # A resoult table 
            sec_row = row_number + 4
            ws['A' + str(sec_row)] = date_from
            ws['A' + str(sec_row + 1)] = date_to
            ws['B' + str(sec_row)] = ws['E18'].value
            ws['B' + str(sec_row + 1)] = str(round(m1_sum, 2)).replace('.', ',')
            ws['C' + str(sec_row)] = ws['G18'].value
            ws['C' + str(sec_row + 1)] = str(round(m2_sum, 2)).replace('.', ',')
            if input_type == '2':
                ws['D' + str(sec_row - 1)] = 'V1, м3'
                ws['D' + str(sec_row)] = ws['D18'].value
                ws['D' + str(sec_row + 1)] = str(round(v1_sum, 2)).replace('.', ',')
                ws['E' + str(sec_row - 1)] = 'V2, м3'
                ws['E' + str(sec_row)] = ws['F18'].value
                ws['E' + str(sec_row + 1)] = str(round(v2_sum, 2)).replace('.', ',')
                ws['F' + str(sec_row - 1)] = 'Q, Гкал'
                ws['F' + str(sec_row)] = ws['J18'].value
                ws['F' + str(sec_row + 1)] = str(round(q_sum, 2)).replace('.', ',')
                for r in range(sec_row - 1, sec_row + 2):
                    thin = Side(border_style="thin", color="000000")
                    ws.cell(r, 7).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(r, 7).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(r, 8).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")
                ws['G' + str(sec_row - 1)] = 'ВНР, час'
                ws['G' + str(sec_row)] = '0'
                ws['G' + str(sec_row + 1)] = str(round(vnr, 2)).replace('.', ',')
                ws['H' + str(sec_row - 1)] = 'ВОС, час'
                ws['H' + str(sec_row)] = '0'
                ws['H' + str(sec_row + 1)] = str(round(vos, 2)).replace('.', ',')
            else:
                ws['D' + str(sec_row)] = ws['J18'].value
                ws['D' + str(sec_row + 1)] = str(round(q_sum, 2)).replace('.', ',')
                ws['E' + str(sec_row)] = '0'
                ws['E' + str(sec_row + 1)] = str(round(vnr, 2)).replace('.', ',')
                ws['F' + str(sec_row)] = '0'
                ws['F' + str(sec_row + 1)] = str(round(vos, 2)).replace('.', ',')
            ws['B6'] = head_data['consumer']
            ws['B7'] = head_data['order']
            ws['B8'] = head_data['adress']
            ws['B11'] = head_data['cold_temp']
            ws['B12'] = head_data['factory_num']
            ws['B13'] = head_data['complex_num']
        curr_dir = self.save_dir + '/Output/' + head_data['save_folder']
        if not os.path.exists(curr_dir):
            os.makedirs(curr_dir)   
        out_type = '_ГВС'
        if input_type == '1':
            out_type = '_отопл'

        name = head_data['consumer'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','') + \
            ' - ' + head_data['adress'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','')
        
        while name in self.report:
            name += '_2'
        template.save(curr_dir + '/' + name + out_type + '.xlsx')
        report_= head_data['save_folder'] + '/' + name + out_type + '.xlsx'+ '\n\n'
        return report_


    def build_xls_by_xls(self, file):
        tv7 = TV7Parser({'ТВ-7': []}, self.my_dir, self.save_dir)
        report = ''
        rep_type = '1'
        if 'ГВС' in file[0]:
            rep_type = '2'

        if file[1]['A1'].value != None:
            if 'Время' in str(file[1]['A2'].value):
                spt_parser = SPTParser([], self.my_dir, self.save_dir)
                columns, gvs_cols = spt_parser.get_columns(list(file[1].rows)[1])
                if '_1' in str(file[1]['A1'].value):
                    report += spt_parser.build_xls(file, columns, '1', 2)
                elif '_2' in str(file[1]['A1'].value):
                    report += spt_parser.build_xls(file, columns, '2', 2)
                else: # '_' not in str(file[1]['A1'].value):
                    report += spt_parser.build_xls(file, columns, '1', 2)
                    report += spt_parser.build_xls(file, columns, '2', 2)

        flag = False
        for cell in file[1][1]:
            if cell.value == None:
                continue
            if 'ОТЧЕТ' in cell.value:
                flag = True
        if flag:
            move_index = 1
            data_indexes = {}
            head_data = {}
            for row in file[1].iter_rows(max_row=14):
                if row[0].value == None: 
                    move_index += 1
                    continue
                if 'ВВОД ' in row[0].value:
                    if row[0].value.split('ВВОД')[1].split('СХЕМА')[0].replace(' ', '') == '2':
                        rep_type = '2'
                if 'Заводской номер' in row[0].value or 'Серийный номер' in row[0].value:
                    head_data = get_head_data(self.my_dir, str(row[0].value).split('Заводской номер ')[1].split(' ')[0], rep_type)
                if 'Дaтa' in row[0].value:
                    data_indexes = tv7.get_columns(row)
                    break

            move_index += 1

            template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx',  read_only=False, data_only=False)
            report, row_inx, out_row_indx, summs, date_from, date_to = tv7.build_xls(file, rep_type, template, data_indexes, head_data, start_read_index=move_index, a_resoul_flag=True, date_format='VKT')
        if '\n' not in report:
            report += '\n'
        return report


    def __call__(self, date_from = '01-01-2023', date_to = '18-01-2023'):
        self.report = '\tВКТ\n' # Window print
        print(len(self.my_parsing_files))
        for file in self.my_parsing_files:
            if '.txt' in file[0]:
                self.report += self.build_xls_by_txt(file).replace(self.save_dir, '')
            elif '.xlsx' in file[0]:
                self.report += self.build_xls_by_xls(file).replace(self.save_dir, '')
            self.report += '\n'

        return self.report