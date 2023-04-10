from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pyexcel
import os
from datetime import datetime
from SPTParser import SPTParser
from HeadData import *


class TMKParser:
    def __init__(self, data_list, curr_dir, save_dir):
        self.sum_report = ''
        self.my_parsing_files = []
        self.my_dir = curr_dir
        self.save_dir = save_dir
        for file in data_list:
            if 'xlsx' not in file:
                if 'xls' in file:
                    pyexcel.save_book_as(file_name=file,
                                dest_file_name=file + 'x')
                    file += 'x'
                    
            if 'xlsx' in file:
                calc_type = None
                ws = load_workbook(file).active
                rep_type = '1'
                if 'ГВС' in str(file.upper()) or 'Тепловая система 2' == str(ws['D9'].value) or 'Тепловой ввод №2' in str(ws['D6'].value):
                    rep_type = '2'

                # If visual type is as SPT
                ws = load_workbook(file).active
                if ws['A1'].value != None:
                    if '_1' in str(ws['A1'].value):
                        calc_type = get_head_data(self.my_dir, str(ws['A1'].value), '1')['type']
                    if '_2' in str(ws['A1'].value):
                        calc_type = get_head_data(self.my_dir, str(ws['A1'].value), '2')['type']

                # Is it really TMK file
                if calc_type is None: 
                    calc_type = get_head_data(self.my_dir, self.get_factory_num(ws[1]), rep_type)['type']

                if calc_type != None:
                    if 'ТМК' in calc_type:
                        self.my_parsing_files.append([file, load_workbook(file).active])
                    else:
                        continue
            else:
                print('TMK wrong file!')
                continue

    # Get str with factory num by row in TMK file
    def get_factory_num(self, row):
        factory_num = ''
        for col in row:
            if col.value == None:
                continue
            if ('ТМК' in col.value or 'Отчет' in col.value) and '№' in col.value:
                factory_num = str(col.value).split('№')[1].replace(' ', '')												
        return factory_num


    def get_columns(self, row, heat_cols={'Время': -1, 't1': -1, 't2': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q': -1, 'Tн': -1,'ВОС': -1, 'НС': -1}):
        ind = 0
        #rep_type = '1'
        for cell in row:
            if cell.value == None:
                ind += 1
                continue
            for key in heat_cols.keys():
                if key in cell.value and 'G1-G2' not in cell.value and 'V1-V2' not in cell.value:
                    heat_cols[key] = ind
            
            if 't3' in cell.value and heat_cols['t1'] == -1:
                heat_cols['t1'] = ind
            if 't4' in cell.value and heat_cols['t2'] == -1:
                heat_cols['t2'] = ind
            if 'V3' in cell.value and heat_cols['V1'] == -1 and 'V1-V2' not in cell.value:
                heat_cols['V1'] = ind
            if 'V4' in cell.value and heat_cols['V2'] == -1 and 'V1-V2' not in cell.value:
                heat_cols['V2'] = ind
            if 'G1' in cell.value and heat_cols['M1'] == -1 and 'G1-G2' not in cell.value:
                heat_cols['M1'] = ind
            if 'G2' in cell.value and heat_cols['M2'] == -1 and 'G1-G2' not in cell.value:
                heat_cols['M2'] = ind
            if 'G3' in cell.value and heat_cols['M1'] == -1 and 'G3-G4' not in cell.value:
                heat_cols['M1'] = ind
            if 'G4' in cell.value and heat_cols['M2'] == -1 and 'G3-G4' not in cell.value:
                heat_cols['M2'] = ind
            if 'Q' in cell.value and heat_cols['Q'] == -1:
                heat_cols['Q'] = ind
            if ('Tр,' in cell.value or 'Время' in cell.value or 'Tраб' in cell.value or 'T1' in cell.value or 'T2' in cell.value) and heat_cols['Tн'] == -1:
                heat_cols['Tн'] = ind
            if 'Tн' in cell.value and heat_cols['Tн'] == -1:
                heat_cols['Tн'] = ind
            if 'НС ТС' in cell.value and heat_cols['НС'] == -1:
                heat_cols['НС'] = ind
            ind += 1

        return heat_cols


    def num_def(self, t, summary_data, data_indexes):
        if summary_data[data_indexes[t]].value != None:
            if '-' not in summary_data[data_indexes[t]].value:
                return round(float(str(summary_data[data_indexes[t]].value).replace(',', '.')), 2)
            else:
                return ' - '
        else:
            return ' - '
        
    def time_conv(self, t, summary_data, data_indexes):
        if summary_data[data_indexes[t]].value != None: 
            if '-' not in str(summary_data[data_indexes[t]].value):
                return round(float(str(summary_data[data_indexes[t]].value).split(':')[0]) + (0.0167) * float(str(summary_data[data_indexes[t]].value).split(':')[1]), 2)
            else:
                return ' - '
        else:
            return ' - '
        

    def build_xls(self, file, rep_type, start_col=1):
        date_from = ''
        date_to = ''
        report = ''
        template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx',  read_only=False, data_only=False)  # Template xlsx file  
        file_name = file[0].split('/')[len(file[0].split('/')) - 1].split('.xlsx')[0]
        
        if file[0].split('/')[len(file[0].split('/')) - 1]:
            template.title = file[0].split('/')[len(file[0].split('/')) - 1]
        ws = template.active
        
        head_data = get_head_data(self.my_dir, self.get_factory_num(file[1][1]), rep_type)
        data_indexes = {'Время': -1, 't1': -1, 't2': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q': -1, 'Tн': -1,'ВОС': -1, 'НС': -1}
        row_index = 1; out_index = 18; data_index = 1
        dm_sum = 0; m1_sum = 0; m2_sum = 0; dv_sum = 0; v1_sum = 0; v2_sum = 0; t1_avg = 0; t2_avg = 0; q_sum = 0; vnr = 0; vos = 0; sum_err = ''    
        
        for row in file[1].iter_rows(min_col=start_col):
            num = '-'
            float_time = '-'
            float_time = lambda t: round(float(str(row[data_indexes[t]].value).split(':')[0]) + 0.0167 * float(str(row[data_indexes[t]].value).split(':')[1]), 2)
            num = lambda t: round(float(str(row[data_indexes[t]].value).replace(',', '.')), 2) if '-' not in row[data_indexes[t]].value else ' - ' if row[data_indexes[t]].value != None else ' - ' 
            st_row = lambda n: str(n).replace('.', ',')
            
            if row_index < 2:
                row_index += 1					
                continue
            
            if row[0].value == 'Дата':
                data_indexes = self.get_columns(row, data_indexes)
                row_index += 1
                data_index = row_index
                continue

            if file[1]['A' + str(row_index-1)].value == 'Дата':
                data_indexes = self.get_columns(row, data_indexes)
                row_index += 1
                data_index = row_index
                continue
            
            if row_index < 9:
                row_index += 1
                data_index = row_index						
                continue

            if row[0].value == None:
                break

            tmp_date = row[0].value.split('.')
            curr_date = datetime.strptime(tmp_date[0] + '-' + tmp_date[1] + '-' + tmp_date[2].split(' ')[0], "%d-%m-%Y").date()
            if date_from == '':
                date_from = curr_date
            date_to = curr_date
            ws.insert_rows(out_index)

            for i in range(1, 14):
                thin = Side(border_style="thin", color="000000")
                ws.cell(out_index, i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(out_index, i).alignment = Alignment(horizontal="center", vertical="center")
            ws['A' + str(out_index)] = str(curr_date.strftime("%d-%m-%Y"))
            if data_indexes['t1'] != -1:
                if num('t1') != ' - ': t1_avg += num('t1') 
                ws['B' + str(out_index)] = st_row(num('t1'))
            if data_indexes['t2'] != -1:
                if num('t2') != ' - ': t2_avg += num('t2')
                ws['C' + str(out_index)] = st_row(num('t2'))
            if data_indexes['V1'] != -1:
                if num('V1') != ' - ': v1_sum += num('V1')
                ws['D' + str(out_index)] = st_row(num('V1'))
                ws['D' + str(out_index + 1)] = str(round(v1_sum, 2)).replace('.', ',')
                if num('V1') != ' - ': ws['H' + str(out_index)] = st_row(round(num('V1'), 2))
                ws['H' + str(out_index + 1)] = str(abs(round(v1_sum, 2))).replace('.', ',')
            if data_indexes['M1'] != -1:
                if num('M1') != ' - ': m1_sum += num('M1')
                ws['E' + str(out_index)] = st_row(num('M1'))
                ws['E' + str(out_index + 1)] = str(round(m1_sum, 2)).replace('.', ',')
                if num('M1') != ' - ': ws['I' + str(out_index)] = st_row(num('M1'))
                ws['I' + str(out_index + 1)] = str(abs(round(m1_sum, 2))).replace('.', ',')
            if data_indexes['V2'] != -1 and num('V2') != ' - ':
                if num('V2') != ' - ': v2_sum += num('V2')
                ws['F' + str(out_index)] = st_row(num('V2'))
                ws['F' + str(out_index + 1)] = str(round(v2_sum, 2)).replace('.', ',')
                if num('V2') != ' - ' and num('V1') != ' - ': ws['H' + str(out_index)] = st_row(round(num('V1') - num('V2'), 2))
                ws['H' + str(out_index + 1)] = str(round(v1_sum - v2_sum, 2)).replace('.', ',')
            if data_indexes['M2'] != -1:
                if num('M2') != ' - ': m2_sum += num('M2')
                ws['G' + str(out_index)] = st_row(num('M2'))
                ws['G' + str(out_index + 1)] = str(round(m2_sum, 2)).replace('.', ',')
                if num('M2') != ' - ' and num('M1') != ' - ': ws['I' + str(out_index)] = st_row(round(num('M1') - num('M2'), 2))
                ws['I' + str(out_index + 1)] = str(round(m1_sum - m2_sum, 2)).replace('.', ',')
            if data_indexes['Q'] != -1:
                if num('Q') != ' - ': q_sum += num('Q')
                ws['J' + str(out_index)] = st_row(num('Q'))
                ws['J' + str(out_index + 1)] = str(round(q_sum, 2)).replace('.', ',')
            if data_indexes['Tн'] != -1:
                if float_time('Tн') != ' - ': vnr += float_time('Tн')
                ws['K' + str(out_index)] = st_row(float_time('Tн'))
                ws['K' + str(out_index + 1)] = str(round(vnr, 2)).replace('.', ',')
            if data_indexes['Tн'] != -1:
                if float_time('Tн') != ' - ': vos += (24.0 - float_time('Tн'))
                ws['L' + str(out_index)] = st_row(round(24.0 - float_time('Tн'), 2))
                ws['L' + str(out_index + 1)] = str(round(vos, 2)).replace('.', ',')

            out_index += 1
            row_index += 1

        ws['B' + str(out_index + 1)] = round(t1_avg/(out_index - 18), 2)
        ws['C' + str(out_index + 1)] = round(t2_avg/(out_index - 18), 2)

        sec_row = out_index + 1
        sum_table_index = row_index + 4

        # Parse resoult table
        while file[1][sum_table_index][0].value != 'Дата':
            sum_table_index += 1
        data_indexes = self.get_columns(file[1][sum_table_index], {'Время': -1, 't1': -1, 't2': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q': -1, 'Tн': -1,'ВОС': -1, 'НС': -1})
        summary_data = file[1][sum_table_index + 1]
        if summary_data[0] == None:
            report += 'Не найдена таблица с итогами в отчете: ' + head_data['factory_num'] + '\n'

        v1_start = 0; v2_start = 0; m1_start = 0; m2_start = 0; q_start = 0; vnr_start = 0; vos_start = 0
        float_time_finnaly = lambda t: self.time_conv(t, summary_data, data_indexes)
        num_finnaly = lambda t: self.num_def(t, summary_data, data_indexes)
        if num_finnaly('M1') != ' - ' and data_indexes['M1'] != -1:
            m1_start = num_finnaly('M1')
            m1_sum += num_finnaly('M1')
        if num_finnaly('M2') != ' - ' and data_indexes['M2'] != -1:
            m2_start = num_finnaly('M2')
            m2_sum += num_finnaly('M2')
        if num_finnaly('V1') != ' - ' and data_indexes['V1'] != -1:
            v1_start = num_finnaly('V1')
            v1_sum += num_finnaly('V1')
        if num_finnaly('V2') != ' - ' and data_indexes['V2'] != -1:
            v2_start = num_finnaly('V2')
            v2_sum += num_finnaly('V2')
        if num_finnaly('Q') != ' - ' and data_indexes['Q'] != -1:
            q_start = num_finnaly('Q')
            q_sum += num_finnaly('Q')
        if float_time_finnaly('Tн') != ' - ' and data_indexes['Tн'] != -1:
            vnr_start = float_time_finnaly('Tн')
            vnr += float_time_finnaly('Tн')
        if float_time_finnaly('ВОС') != ' - ' and data_indexes['ВОС'] != -1:
            vnr_start = float_time_finnaly('ВОС')
            vnr += float_time_finnaly('ВОС')


        sec_row += 3
        # A resoult table 
        ws['A' + str(sec_row)] = date_from  
        ws['A' + str(sec_row + 1)] = date_to
        ws['B' + str(sec_row)] = str(round(m1_start, 2)).replace('.', ',')
        ws['B' + str(sec_row + 1)] = str(round(m1_sum, 2)).replace('.', ',')
        ws['C' + str(sec_row)] = str(round(m2_start, 2)).replace('.', ',')
        ws['C' + str(sec_row + 1)] = str(round(m2_sum, 2)).replace('.', ',')
        q_col = 'D'; vnr_col = 'E'; vos_col = 'F'
        if rep_type == '2':
            ws['D' + str(sec_row - 1)] = 'V1, м3'
            ws['D' + str(sec_row)] = str(round(v1_start, 2)).replace('.', ',')
            ws['D' + str(sec_row + 1)] = str(round(v1_sum, 2)).replace('.', ',')
            ws['E' + str(sec_row - 1)] = 'V2, м3'
            ws['E' + str(sec_row)] = str(round(v2_start, 2)).replace('.', ',')
            ws['E' + str(sec_row + 1)] = str(round(v2_sum, 2)).replace('.', ',')
            for r in range(sec_row - 1, sec_row + 2):
                thin = Side(border_style="thin", color="000000")
                ws.cell(r, 7).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(r, 7).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, 8).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")

            q_col = 'F'; vnr_col = 'G'; vos_col = 'H'
            ws['F' + str(sec_row - 1)] = 'Q, Гкал'
            ws['G' + str(sec_row - 1)] = 'ВНР, час'
            ws['H' + str(sec_row - 1)] = 'ВОС, час'
            
        ws[q_col + str(sec_row)] = str(round(q_start, 2)).replace('.', ',')
        ws[q_col + str(sec_row + 1)] = str(round(q_sum, 2)).replace('.', ',')
        ws[vnr_col + str(sec_row)] = str(round(vnr_start, 2)).replace('.', ',')
        ws[vnr_col + str(sec_row + 1)] = str(round(vnr, 2)).replace('.', ',')
        ws[vos_col + str(sec_row)] = str(round(vos_start, 2)).replace('.', ',')
        ws[vos_col + str(sec_row + 1)] = str(round(vos, 2)).replace('.', ',')

        # Fill head data
        ws['A1'] = str(ws['A1'].value).replace('май', get_month(datetime.strftime(date_to, "%d-%m-%Y")))
        ws['B3'] = datetime.strftime(date_from, "%d-%m-%Y")
        ws['C3'] = datetime.strftime(date_to, "%d-%m-%Y")
        ws['B4'] = datetime.now().strftime("%d-%m-%Y")
        ws['B5'] = rep_type
        ws['B6'] = head_data['consumer']
        ws['B7'] = head_data['order']
        ws['B8'] = head_data['adress']
        ws['B11'] = head_data['cold_temp']
        ws['B12'] = head_data['factory_num'].replace('_2', '').replace('_1', '')
        ws['B13'] = head_data['complex_num']
        curr_dir = self.save_dir + '/Output/' + head_data['save_folder']
        if not os.path.exists(curr_dir):
            os.makedirs(curr_dir)

        str_rep = '_ГВС'
        if rep_type == '1':
            str_rep = '_отопл'

        name = head_data['consumer'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','') + \
            ' - ' + head_data['adress'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','')
        while name in self.sum_report:
            name += '_2'
        template.save(curr_dir + '/' + name + str_rep + '.xlsx')
        report += head_data['save_folder'] + '/' + name + str_rep +'.xlsx'

        return report + '\n'


    def __call__(self):
        self.sum_report = '\tТМК\n'
        print(len(self.my_parsing_files))
        for file in self.my_parsing_files:
            if file[1]['A1'].value != None:
                if 'Время' in str(file[1]['A2'].value):
                    spt_parser = SPTParser([], self.my_dir, self.save_dir)
                    columns, gvs_cols = spt_parser.get_columns(list(file[1].rows)[1])
                    if '_1' in str(file[1]['A1'].value):
                        self.sum_report += spt_parser.build_xls(file, columns, '1', 2)
                        continue
                    elif '_2' in str(file[1]['A1'].value):
                        self.sum_report += spt_parser.build_xls(file, columns, '2', 2)
                        continue
                    else: # '_' not in str(file[1]['A1'].value):
                        self.sum_report += spt_parser.build_xls(file, columns, '1', 2)
                        self.sum_report += spt_parser.build_xls(file, columns, '2', 2)
                        continue

            rep_type = '1'
            if 'ГВС' in str(file[0].upper()) or 'Тепловая система 2' == str(file[1]['D9'].value) or 'Тепловой ввод №2' in str(file[1]['D6'].value):
                rep_type = '2'

            self.sum_report += self.build_xls(file, rep_type) + '\n' 

        return self.sum_report