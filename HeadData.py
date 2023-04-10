from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pyexcel
import os
from datetime import datetime


months = {'01':'январь', '02':'февраль', '03':'март', '04':'апрель',\
         '05':'май', '06':'июнь', '07':'июль', '08':'август',\
         '09':'сеньтябрь', '10':'октябрь', '11':'ноябрь', '12':'декабрь'}


def get_month(m='01-01-2023'):
    return months[m[3:5]]

    
def get_head_data(my_dir, factory_num, inp_type):
    header_ws = load_workbook(my_dir + '\Templates\HeadDat.xlsx', data_only=False).active
    head_data = {'factory_num': '', 'complex_num': '', 'consumer': '', 'order': '', 'adress': '', 'cold_temp': '5,0', 'save_folder': '', 'type': ''}
    for i in range(2, 426):
        if  str(header_ws['A' + str(i)].value) in factory_num + '_' + inp_type:
            head_data['factory_num'] = str(header_ws['A' + str(i)].value).replace('_1', '').replace('_2', '')
            head_data['complex_num'] = header_ws['C' + str(i)].value
            head_data['consumer'] = header_ws['D' + str(i)].value
            head_data['order'] = header_ws['E' + str(i)].value
            head_data['adress'] = header_ws['F' + str(i)].value
            head_data['cold_temp'] = header_ws['H' + str(i)].value
            head_data['save_folder'] = header_ws['K' + str(i)].value
            head_data['type'] = header_ws['B' + str(i)].value
        elif '-' in factory_num and str(header_ws['A' + str(i)].value).split('_')[0] in factory_num + '_' + inp_type:
            head_data['factory_num'] = factory_num
            head_data['complex_num'] = header_ws['C' + str(i)].value
            head_data['consumer'] = header_ws['D' + str(i)].value
            head_data['order'] = header_ws['E' + str(i)].value
            head_data['adress'] = header_ws['F' + str(i)].value
            head_data['cold_temp'] = header_ws['H' + str(i)].value
            head_data['save_folder'] = header_ws['K' + str(i)].value
            head_data['type'] = header_ws['B' + str(i)].value

    return head_data