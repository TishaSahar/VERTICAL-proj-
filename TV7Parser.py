from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pyexcel
import os
from datetime import datetime
from HeadData import *


def err_vkt(ns_indx, row):
    err = ''
    for cell in row[ns_indx: ns_indx + 8]:
        if cell.value is None:
            continue
        if '4' in str(cell.value) or '6' in str(cell.value):
            err = '1'

    return err


def err_tv7(ns_indx, row):
    err = ''
    for cell in row[1: ns_indx + 8]:
        c = cell.value
        if c is None:
            continue
        c = str(c)
        if '<' in str(c) and '3, 4' not in err:
            err += '3, 4'
        if '>' in str(c) and '5' not in err:
            err += '5'
        if '!' in str(c) and '1' not in err: 
            err += '1'

    return err


class TV7Parser:
    def __init__(self, data_list, curr_dir, save_dir):
        self.summary_rep = ''
        self.my_parsing_files = []
        self.my_dir = curr_dir
        self.save_dir = save_dir

        if len(data_list) != 0:
            for file in data_list:
                head_data = {'factory_num': '', 'complex_num': '', 'consumer': '', 'order': '', 'adress': '', 'cold_temp': '5,0', 'save_folder': '', 'type': ''}
                if 'xls' in file:
                    if 'xlsx' not in file:
                        pyexcel.save_book_as(file_name=file,
                                    dest_file_name=file + 'x')
                        file += 'x'

                    rep_type = '1'
                    if 'ГВС' in file[0].upper():
                        rep_type = '2'
                    ws = load_workbook(file).active
                    for row in ws.iter_rows(max_row=16):
                        if row[0].value == None: 
                            continue
                        if 'Серийный номер' in str(row[0].value):
                            if 'ТВ2' in row[0].value:
                                rep_type = '2'
                            if 'ТВ1' in row[0].value:
                                rep_type = '1'
                            head_data = get_head_data(self.my_dir, str(row[0].value).split('Серийный номер ')[1].split(',')[0], rep_type)
                    if head_data['type'] != None:
                        if 'ТВ-7' in str(head_data['type']).replace(' ', ''):
                            self.my_parsing_files.append([file, load_workbook(file).active])
                        else:
                            continue
                    else:
                        print('Can not to expect type in: ',  str(file))
                else:
                    continue


    def get_columns(self, row):
        heat_cols = {'Время': -1, 't1': -1, 't2': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q': -1, 'ВНР': -1,'ВОС': -1, 'НС': -1}
        ind = 0
        for cell in row:
            if cell.value == None:
                ind += 1
                continue

            for key in heat_cols.keys():
                if key in cell.value and heat_cols[key] == -1:
                    heat_cols[key] = ind

            if 't3' in cell.value and heat_cols['t1'] == -1:
                heat_cols['t1'] = ind
            if 't4' in cell.value and heat_cols['t2'] == -1:
                heat_cols['t2'] = ind

            if 'V3' in cell.value and heat_cols['V1'] == -1 and 'dV' not in cell.value:
                heat_cols['V1'] = ind
            if 'V4' in cell.value and heat_cols['V2'] == -1 and 'dV' not in cell.value:
                heat_cols['V2'] = ind

            if 'M3' in cell.value and heat_cols['M1'] == -1 and 'dM' not in cell.value:
                heat_cols['M1'] = ind
            if 'M4' in cell.value and heat_cols['M2'] == -1 and 'dM' not in cell.value:
                heat_cols['M2'] = ind

            if 'H.C.' in cell.value:
                heat_cols['НС'] = ind
            if 'BНP' in cell.value:
                heat_cols['ВНР'] = ind
            if 'BOC' in cell.value:
                heat_cols['ВОС'] = ind
            ind += 1

        return heat_cols

    def num_from_data(self, t, summary_data, data_indexes):
        if summary_data[data_indexes[t]].value != None:
            if '-' not in str(summary_data[data_indexes[t]].value):
                return round(float(summary_data[data_indexes[t]].value), 2)
            else:
                return ' - '
        else: 
            return ' - '

    def build_xls(self, file, rep_type,  template, data_indexes=[], head_data={}, start_read_index=1, start_out_index=18, summs=[0,0,0,\
                                                                                                                                                0,0,0,\
                                                                                                                                                0,0,0, ''],a_resoul_flag=False, date_format='TV7', date_from = '',date_to = ''):
        report = '' # Out text, will be printed into my textBrowser
        if file[0].split('/')[len(file[0].split('/')) - 1]:
            template.title = file[0].split('/')[len(file[0].split('/')) - 1]
        
        ws = template.active
        row_index = start_read_index; out_index = start_out_index
        t1_avg, t2_avg, m1_sum, m2_sum, v1_sum, v2_sum, q_sum, vnr, vos, sum_err = summs    

        for row in file[1].iter_rows(min_row=row_index):
            num = lambda t: round(float(row[data_indexes[t]].value), 2) if row[data_indexes[t]].value != None else ' - '
            num3 = lambda t: round(float(row[data_indexes[t]].value), 3) if row[data_indexes[t]].value != None else ' - '
            
            st_row = lambda n: str(n).replace('.', ',')
            curr_date = ''
            if 'Итого:' == row[0].value or 'Итого/Средн' == row[0].value:
                break

            if row[0].value == None:
                row_index += 1
                continue
            
            tmp_date = str(row[0].value).replace('-', '.').replace('00:00:00', '').split('.')
            if len(tmp_date) < 3 or len(str(tmp_date[1])) != 2:
                row_index += 1
                continue
            
            # Parse dates
            if date_format == 'VKT':
                curr_date = datetime.strptime(tmp_date[2].replace(' ', '') + '-' + tmp_date[1].replace(' ', '') + '-' + tmp_date[0].split(' ')[0].replace(' ', ''), "%d-%m-%Y").date()
            else:
                curr_date = datetime.strptime(tmp_date[0] + '-' + tmp_date[1] + '-' + '20' + tmp_date[2].split(' ')[0], "%d-%m-%Y").date()
            
            # Fill the start and end dates
            if date_from == '':
                date_from = curr_date
            if str(curr_date) != '':
                date_to = curr_date

            # Fill main table
            ws.insert_rows(out_index)
            for i in range(1, 14):
                thin = Side(border_style="thin", color="000000")
                ws.cell(out_index, i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(out_index, i).alignment = Alignment(horizontal="center", vertical="center")
            ws['A' + str(out_index)] = str(curr_date.strftime("%d-%m-%Y"))
            if data_indexes['t1'] != -1:
                if num('t1') != ' - ': t1_avg += num('t1') 
                ws['B' + str(out_index)] = st_row(num('t1'))
            if data_indexes['t2'] != -1:
                if num('t2') != ' - ': t2_avg += num('t2')
                ws['C' + str(out_index)] = st_row(num('t2'))
            if data_indexes['V1'] != -1:
                if num('V1') != ' - ': v1_sum += num('V1')
                ws['D' + str(out_index)] = st_row(num('V1'))
                ws['D' + str(out_index + 1)] = str(round(v1_sum, 2)).replace('.', ',')
                ws['H' + str(out_index)] = st_row(round(num('V1'), 2))
                ws['H' + str(out_index + 1)] = str(abs(round(v1_sum, 3))).replace('.', ',')
            if data_indexes['M1'] != -1:
                if num('M1') != ' - ': m1_sum += num('M1')
                ws['E' + str(out_index)] = st_row(num('M1'))
                ws['E' + str(out_index + 1)] = str(round(m1_sum, 2)).replace('.', ',')
                ws['I' + str(out_index)] = st_row(num('M1'))
                ws['I' + str(out_index + 1)] = str(abs(round(m1_sum, 3))).replace('.', ',')
            if data_indexes['V2'] != -1 and num('V2') != ' - ':
                v2_sum += num('V2')
                ws['F' + str(out_index)] = st_row(num('V2'))
                ws['F' + str(out_index + 1)] = str(round(v2_sum, 3)).replace('.', ',')
                ws['H' + str(out_index)] = st_row(round(num('V1') - num('V2'), 2))
                ws['H' + str(out_index + 1)] = str(round(v1_sum - v2_sum, 3)).replace('.', ',')
            if data_indexes['M2'] != -1:
                if num('M2') != ' - ': m2_sum += num('M2')
                ws['G' + str(out_index)] = st_row(num('M2'))
                ws['G' + str(out_index + 1)] = str(round(m2_sum, 3)).replace('.', ',')
                ws['I' + str(out_index)] = st_row(round(num('M1') - num('M2'), 2))
                ws['I' + str(out_index + 1)] = str(round(m1_sum - m2_sum, 3)).replace('.', ',')
            if data_indexes['Q'] != -1:
                if num('Q') != ' - ': q_sum += num3('Q')
                ws['J' + str(out_index)] = st_row(num3('Q'))
                ws['J' + str(out_index + 1)] = str(round(q_sum, 3)).replace('.', ',')
            if data_indexes['ВНР'] != -1:
                if num('ВНР') != ' - ': vnr += num('ВНР')
                ws['K' + str(out_index)] = st_row(num('ВНР'))
                ws['K' + str(out_index + 1)] = str(round(vnr, 2)).replace('.', ',')
            if data_indexes['ВНР'] != -1:
                if num('ВНР') != ' - ': vos += (24.0 - num('ВНР'))
                ws['L' + str(out_index)] = st_row(24.0 - num('ВНР'))
                ws['L' + str(out_index + 1)] = str(round(vos, 2)).replace('.', ',')
            # If we are have some errors from this date
            if data_indexes['НС'] != -1:
                err = ''
                if date_format == 'VKT':
                    err = err_vkt(data_indexes['НС'], row)
                else:
                    err = err_tv7(data_indexes['НС'], row)

                ws['M' + str(out_index)] = err
                if err not in sum_err:
                    if sum_err == '':
                        sum_err += err
                    else:
                        for ch in err:
                            if ch not in sum_err:
                                sum_err += ',' + ch

                ws['M' + str(out_index + 1)] = sum_err
            out_index += 1
            row_index += 1

        if a_resoul_flag == True:
            ws['B' + str(out_index + 1)] = round(t1_avg/(out_index - 18), 2)
            ws['C' + str(out_index + 1)] = round(t2_avg/(out_index - 18), 2)

            sec_row = out_index + 1
            # Parse resoult table
            summary_data = file[1][row_index + 8]
            if date_format == 'VKT':
                row_shift = 0
                for row in file[1].iter_rows(min_row=row_index):
                    if row[0].value == None:
                        row_shift += 1
                        continue
                    if 'Дaтa' in row[0].value:
                        break
                    row_shift += 1

                data_indexes = self.get_columns(file[1][row_index + row_shift])
                summary_data = file[1][row_index + row_shift + 3]
                if summary_data[0].value == None:
                    if file[1][row_index + row_shift + 2][0].value != None:
                        summary_data = file[1][row_index + row_shift + 2]
                    else:
                        report += 'Не найдена таблица с итогами в отчете ВКТ: ' + head_data['factory_num'] + '\n'
            else:
                data_indexes = self.get_columns(file[1][row_index + 3])
                summary_data = file[1][row_index + 9]
            
            v1_start = 0; v2_start = 0; m1_start = 0; m2_start = 0; q_start = 0; vnr_start = 0; vos_start = 0
            num_finnaly = lambda t: self.num_from_data(t, summary_data, data_indexes)
            if num_finnaly('M1') != ' - ' and data_indexes['M1'] != -1:
                m1_start = num_finnaly('M1') - m1_sum
                m1_sum = num_finnaly('M1')
            if num_finnaly('M2') != ' - ' and data_indexes['M2'] != -1:
                m2_start = num_finnaly('M2') - m2_sum
                m2_sum = num_finnaly('M2')
            if num_finnaly('V1') != ' - ' and data_indexes['V1'] != -1:
                v1_start = num_finnaly('V1') - v1_sum
                v1_sum = num_finnaly('V1')
            if num_finnaly('V2') != ' - ' and data_indexes['V2'] != -1:
                v2_start = num_finnaly('V2') - v2_sum
                v2_sum = num_finnaly('V2')
            if num_finnaly('Q') != ' - ' and data_indexes['Q'] != -1:
                q_start = num_finnaly('Q') - q_sum
                q_sum = num_finnaly('Q')
            if num_finnaly('ВНР') != ' - ' and data_indexes['ВНР'] != -1:
                vnr_start = num_finnaly('ВНР') - vnr
                vnr = num_finnaly('ВНР')
            if num_finnaly('ВОС') != ' - ' and data_indexes['ВОС'] != -1:
                vos_start = num_finnaly('ВОС') - vos
                vos = num_finnaly('ВОС')

            sec_row += 3
            # Fill resoult table   
            ws['A' + str(sec_row)] = date_from  
            ws['A' + str(sec_row + 1)] = date_to
            ws['B' + str(sec_row)] = str(round(m1_start, 2)).replace('.', ',')
            ws['B' + str(sec_row + 1)] = str(round(m1_sum, 2)).replace('.', ',')
            ws['C' + str(sec_row)] = str(round(m2_start, 2)).replace('.', ',')
            ws['C' + str(sec_row + 1)] = str(round(m2_sum, 2)).replace('.', ',')
            q_col = 'D'; vnr_col = 'E'; vos_col = 'F'
            if rep_type == '2':
                ws['D' + str(sec_row - 1)] = 'V1, м3'
                ws['D' + str(sec_row)] = str(round(v1_start, 2)).replace('.', ',')
                ws['D' + str(sec_row + 1)] = str(round(v1_sum, 2)).replace('.', ',')
                ws['E' + str(sec_row - 1)] = 'V2, м3'
                ws['E' + str(sec_row)] = str(round(v2_start, 2)).replace('.', ',')
                ws['E' + str(sec_row + 1)] = str(round(v2_sum, 2)).replace('.', ',')
                for r in range(sec_row - 1, sec_row + 2):
                    thin = Side(border_style="thin", color="000000")
                    ws.cell(r, 7).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(r, 7).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(r, 8).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")
                q_col = 'F'; vnr_col = 'G'; vos_col = 'H'
                ws['F' + str(sec_row - 1)] = 'Q, Гкал'
                ws['G' + str(sec_row - 1)] = 'ВНР, час'
                ws['H' + str(sec_row - 1)] = 'ВОС, час'

            ws[q_col + str(sec_row)] = str(round(q_start, 2)).replace('.', ',')
            ws[q_col + str(sec_row + 1)] = str(round(q_sum, 2)).replace('.', ',')
            ws[vnr_col + str(sec_row)] = str(round(vnr_start, 2)).replace('.', ',')
            ws[vnr_col + str(sec_row + 1)] = str(round(vnr, 2)).replace('.', ',')
            ws[vos_col + str(sec_row)] = str(round(vos_start, 2)).replace('.', ',')
            ws[vos_col + str(sec_row + 1)] = str(round(vos, 2)).replace('.', ',')

        # Fill head data
        ws['A1'] = str(ws['A1'].value).replace('май', get_month(str(date_to)[8:10] + '-' + str(date_to)[5:7] + '-' + str(date_to)[0:4]))
        ws['B3'] = date_from
        ws['C3'] = date_to
        ws['B4'] = datetime.now().strftime("%d-%m-%Y")
        ws['B5'] = str(rep_type)
        ws['B6'] = head_data['consumer']
        ws['B7'] = head_data['order']
        ws['B8'] = head_data['adress']
        ws['B11'] = head_data['cold_temp']
        ws['B12'] = head_data['factory_num'].replace('_2', '').replace('_1', '')
        ws['B13'] = head_data['complex_num']
        curr_dir = self.save_dir + '/Output/' + head_data['save_folder']
        if not os.path.exists(curr_dir):
            os.makedirs(curr_dir)
        str_rep = '_ГВС'
        if rep_type == '1':
            str_rep = '_отопл'
        
        name = head_data['consumer'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','') + \
            ' - ' + head_data['adress'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','')

        while name in self.summary_rep:
            name += '_2'

        template.save(curr_dir + '/' + name + str_rep + '.xlsx')
        report += curr_dir + '/' + name + str_rep +'.xlsx'
        return [report, row_index, out_index, [t1_avg, t2_avg, m1_sum, m2_sum, v1_sum, v2_sum, q_sum, vnr, vos, sum_err], date_from, date_to]


    def __call__(self):
        self.summary_rep = '\tТВ - 7\n'
        print(len(self.my_parsing_files))
        for file in self.my_parsing_files:
            rep_type = '1'
            if 'ГВС' in file[0].upper():
                rep_type = '2'
            move_index = 1
            count_of_tables = 0
            for row in file[1].iter_rows():
                if row[0].value == None:
                    continue
                if 'ОТЧЕТ' in str(row[0].value):
                    count_of_tables += 1

            for row in file[1].iter_rows(max_row=16):
                if row[0].value == None: 
                    move_index += 1
                    continue
                if 'Серийный номер' in row[0].value:
                    if 'ТВ2' in row[0].value:
                        rep_type = '2'
                    if 'ТВ1' in row[0].value:
                        rep_type = '1'
                    head_data = get_head_data(self.my_dir, str(row[0].value).split('Серийный номер ')[1].split(',')[0], rep_type)
                if 'Дата/время' in row[0].value:
                    data_indexes = self.get_columns(row)
                    break
                move_index += 1

            if count_of_tables == 2:
                template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx',  read_only=False, data_only=False)
                report, row_inx, out_row_indx, summs, date_from, date_to = self.build_xls(file, rep_type, template, data_indexes, head_data, start_read_index=move_index, a_resoul_flag=False)
                move_index = 10
                for row in file[1].iter_rows(min_row=row_inx+10):
                    if row[0].value == None: 
                        move_index += 1
                        continue
                    if 'Серийный номер' in row[0].value:
                        head_data = get_head_data(self.my_dir, str(row[0].value).split('Серийный номер ')[1].split(',')[0], rep_type)
                        if 'ТВ2' in row[0].value:
                            rep_type = '2'
                        if 'ТВ1' in row[0].value:
                            rep_type = '1'
                    if 'Дата/время' in row[0].value:
                        data_indexes = self.get_columns(row)
                        break
                    move_index += 1

                data_indexes = self.get_columns(file[1][row_inx+move_index])
                template = load_workbook(report,  read_only=False, data_only=False)
                report, row_inx, out_row_indx, summs, date_from, date_to = self.build_xls(file, rep_type, template, data_indexes, head_data, row_inx + move_index, out_row_indx, summs, True, date_from=date_from, date_to=date_to)
            else:
                template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx',  read_only=False, data_only=False)
                report, row_inx, out_row_indx, summs, date_from, date_to = self.build_xls(file, rep_type, template, data_indexes, head_data, start_read_index=move_index, a_resoul_flag=True)

            self.summary_rep += '\n' + report.replace(self.save_dir, '') + '\n'
        return self.summary_rep