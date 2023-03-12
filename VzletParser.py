from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pyexcel
import os
from datetime import datetime

months = {'01':'январь', '02':'февраль', '03':'март', '04':'апрель',\
         '05':'май', '06':'июнь', '07':'июль', '08':'август',\
         '09':'сеньтябрь', '10':'октябрь', '11':'ноябрь', '12':'декабрь'}
def get_month(m='01-01-2023'):
    return months[m[3:5]]


def get_head_data(my_dir, factory_num, inp_type):
    header_ws = load_workbook(my_dir + '\Templates\HeadDat.xlsx', data_only=False).active
    head_data = {'factory_num': '', 'complex_num': '', 'consumer': '', 'order': '', 'adress': '', 'cold_temp': '5,0', 'save_folder': ''}
    for i in range(2, 426):
        if  str(header_ws['A' + str(i)].value) in factory_num + '_' + inp_type:
            head_data['factory_num'] = header_ws['A' + str(i)].value.replace('_1', '').replace('_2', '')
            head_data['complex_num'] = header_ws['C' + str(i)].value
            head_data['consumer'] = header_ws['D' + str(i)].value
            head_data['order'] = header_ws['E' + str(i)].value
            head_data['adress'] = header_ws['F' + str(i)].value
            head_data['cold_temp'] = header_ws['H' + str(i)].value
            head_data['save_folder'] = header_ws['K' + str(i)].value
        elif '-' in factory_num and str(header_ws['A' + str(i)].value.split('_')[0]) in factory_num + '_' + inp_type:
            head_data['factory_num'] = factory_num
            head_data['complex_num'] = header_ws['C' + str(i)].value
            head_data['consumer'] = header_ws['D' + str(i)].value
            head_data['order'] = header_ws['E' + str(i)].value
            head_data['adress'] = header_ws['F' + str(i)].value
            head_data['cold_temp'] = header_ws['H' + str(i)].value
            head_data['save_folder'] = header_ws['K' + str(i)].value
    
    return head_data


class VzletParser:
    def __init__(self, data_list, curr_dir, save_dir):
        self.my_parsing_files = []
        self.my_dir = curr_dir
        self.save_dir = save_dir
        for file in data_list['ВЗЛЕТ']:
            if 'xlsx' not in file:
                if 'xls' in file:
                    pyexcel.save_book_as(file_name=file,
                                dest_file_name=file + 'x')
                    file += 'x'
                else:
                    print('Vzlet wrong file!')
                    continue

            self.my_parsing_files.append([file, load_workbook(file).active])

    def get_columns(self, row):
        heat_cols = {'Время': -1, 't1': -1, 't2': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q': -1, 'Т, ч': -1}
        ind = 0
        for cell in row:
            if cell.value == None:
                ind += 1
                continue
            for key in heat_cols.keys():
                if key in cell.value:
                    heat_cols[key] = ind
            if 'T1' in cell.value:
                heat_cols['t1'] = ind
            if 'T2' in cell.value:
                heat_cols['t2'] = ind
            if 'V3' in cell.value:
                heat_cols['V1'] = ind
            if 'V4' in cell.value:
                heat_cols['V2'] = ind
            if 'М1, т' in cell.value or 'M3' in cell.value:
                heat_cols['M1'] = ind
            if 'M2, т' in cell.value or 'M4' in cell.value:
                heat_cols['M2'] = ind
            if 'Нараб' in cell.value or 'Время раб' in cell.value:
                heat_cols['Т, ч'] = ind
            ind += 1
        return heat_cols

     
    def build_xls(self, file, rep_type, date_from = '01-01-2023', date_to = '18-01-2023'):
        report = ''
        template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx',  read_only=False, data_only=False)  # Template xlsx file  
        file_name = file[0].split('/')[len(file[0].split('/')) - 1].split('.xlsx')[0]
        if file[0].split('/')[len(file[0].split('/')) - 1]:
            template.title = file[0].split('/')[len(file[0].split('/')) - 1]
        ws = template.active

        head_data = {}; data_indexes = {}
        row_index = 1; out_index = 18
        m1_sum = 0; m2_sum = 0; v1_sum = 0; v2_sum = 0; t1_avg = 0; t2_avg = 0; q_sum = 0; vnr = 0; vos = 0; sum_err = ''    
        for row in file[1].iter_rows():
            num = lambda t: round(float(row[data_indexes[t]].value), 2) if row[data_indexes[t]].value != None else ' - '
            st_row = lambda n: str(n).replace('.', ',')
            if row_index == 1:
                row_index += 1
                head_data = get_head_data(self.my_dir, row[0].value, rep_type)
                continue
            if row_index == 2:
                row_index += 1
                data_indexes = self.get_columns(row)
                continue
            if row[0].value == None:
                break
            
            curr_date = datetime.date(row[0].value)
            if (curr_date >= datetime.strptime(date_from, "%d-%m-%Y").date() and\
                curr_date <= datetime.strptime(date_to, "%d-%m-%Y").date()):
                ws.insert_rows(out_index)
                for i in range(1, 14):
                    thin = Side(border_style="thin", color="000000")
                    ws.cell(out_index, i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(out_index, i).alignment = Alignment(horizontal="center", vertical="center")
                ws['A' + str(out_index)] = str(curr_date.strftime("%d-%m-%Y"))
                if data_indexes['t1'] != -1:
                    if num('t1') != ' - ': t1_avg += num('t1') 
                    ws['B' + str(out_index)] = st_row(num('t1'))
                if data_indexes['t2'] != -1:
                    if num('t2') != ' - ': t2_avg += num('t2')
                    ws['C' + str(out_index)] = st_row(num('t2'))
                if data_indexes['V1'] != -1:
                    if num('V1') != ' - ': v1_sum += num('V1')
                    ws['D' + str(out_index)] = st_row(num('V1'))
                    ws['D' + str(out_index + 1)] = str(round(v1_sum, 2)).replace('.', ',')
                    ws['H' + str(out_index)] = st_row(round(num('V1'), 2))
                    ws['H' + str(out_index + 1)] = str(abs(round(v1_sum, 2))).replace('.', ',')
                if data_indexes['M1'] != -1:
                    if num('M1') != ' - ': m1_sum += num('M1')
                    ws['E' + str(out_index)] = st_row(num('M1'))
                    ws['E' + str(out_index + 1)] = str(round(m1_sum, 2)).replace('.', ',')
                    ws['I' + str(out_index)] = st_row(num('M1'))
                    ws['I' + str(out_index + 1)] = str(abs(round(m1_sum, 2))).replace('.', ',')
                if data_indexes['V2'] != -1 and num('V2') != ' - ':
                    v2_sum += num('V2')
                    ws['F' + str(out_index)] = st_row(num('V2'))
                    ws['F' + str(out_index + 1)] = str(round(v2_sum, 2)).replace('.', ',')
                    if num('V1') != ' - ': ws['H' + str(out_index)] = st_row(abs(round(num('V2') - num('V1'), 2))) 
                    ws['H' + str(out_index + 1)] = str(abs(round(v2_sum - v1_sum, 2))).replace('.', ',')
                if data_indexes['M2'] != -1 and num('M2') != ' - ':
                    m2_sum += num('M2')
                    ws['G' + str(out_index)] = st_row(num('M2'))
                    ws['G' + str(out_index + 1)] = str(round(m2_sum, 2)).replace('.', ',')
                    if num('M1') != ' - ': ws['I' + str(out_index)] = st_row(abs(round(num('M2') - num('M1'), 2))) 
                    ws['I' + str(out_index + 1)] = str(abs(round(m2_sum - m1_sum, 2))).replace('.', ',')
                if data_indexes['Q'] != -1 and num('Q') != ' - ':
                    q_sum += num('Q')
                    ws['J' + str(out_index)] = st_row(num('Q'))
                    ws['J' + str(out_index + 1)] = str(round(q_sum, 2)).replace('.', ',')
                if data_indexes['Т, ч'] != -1 and num('Т, ч') != ' - ': 
                    vnr += num('Т, ч')
                    ws['K' + str(out_index)] = st_row(num('Т, ч'))
                    ws['K' + str(out_index + 1)] = str(round(vnr, 2)).replace('.', ',')
                    vos += (24.0 - num('Т, ч'))
                    ws['L' + str(out_index)] = st_row(round(24.0 - num('Т, ч'), 2))
                    ws['L' + str(out_index + 1)] = str(round(vos, 2)).replace('.', ',')
                else:
                    vnr += 24
                    ws['K' + str(out_index)] = st_row(24.0)
                    ws['K' + str(out_index + 1)] = str(round(vnr, 2)).replace('.', ',')
                    vos += 0
                    ws['L' + str(out_index)] = st_row(0)
                    ws['L' + str(out_index + 1)] = str(round(vos, 2)).replace('.', ',')
                out_index += 1
            row_index += 1

        ws['B' + str(out_index + 1)] = round(t1_avg/(out_index - 17), 2)
        ws['C' + str(out_index + 1)] = round(t2_avg/(out_index - 17), 2)

        sec_row = out_index + 1
        sum_table_index = row_index + 1
        v1_start = 0; v2_start = 0; m1_start = 0; m2_start = 0; q_start = 0; vnr_start = 0; vos_start = 0
        if None != file[1][sum_table_index][0].value:
            if 'Дата и время' in file[1][sum_table_index][0].value:
                data_indexes = self.get_columns(file[1][sum_table_index])
        else:
            data_indexes = {'Время': -1, 't1': -1, 't2': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q': -1, 'Т, ч': -1}
        
        summary_data = file[1][sum_table_index + 1]
        num_finnaly = lambda t: round(float(str(summary_data[data_indexes[t]].value).replace(',', '.')), 2) if summary_data[data_indexes[t]].value != None else ' - '
        
        if summary_data[data_indexes['M1']].value != None and data_indexes['M1'] != -1:
            m1_start = num_finnaly('M1')
            m1_sum += num_finnaly('M1')
        if summary_data[data_indexes['M2']].value != None and data_indexes['M2'] != -1:
            m2_start = num_finnaly('M2')
            m2_sum += num_finnaly('M2')
        if summary_data[data_indexes['V1']].value != None and data_indexes['V1'] != -1:
            v1_start = num_finnaly('V1')
            v1_sum += num_finnaly('V1')
        if summary_data[data_indexes['V2']].value != None and data_indexes['V2'] != -1:
            v2_start = num_finnaly('V2')
            v2_sum += num_finnaly('V2')
        if summary_data[data_indexes['Q']].value != None and data_indexes['Q'] != -1:
            q_start = num_finnaly('Q')
            q_sum += num_finnaly('Q')
        if summary_data[data_indexes['Т, ч']].value != None and data_indexes['Т, ч'] != -1:
            vnr_start = num_finnaly('Т, ч')
            vnr += num_finnaly('Т, ч')

        sec_row += 3
        # A resoult table 
        ws['A' + str(sec_row)] = date_from  
        ws['A' + str(sec_row + 1)] = date_to
        ws['B' + str(sec_row)] = str(round(m1_start, 2)).replace('.', ',')
        ws['B' + str(sec_row + 1)] = str(round(m1_sum, 2)).replace('.', ',')
        ws['C' + str(sec_row)] = str(round(m2_start, 2)).replace('.', ',')
        ws['C' + str(sec_row + 1)] = str(round(m2_sum, 2)).replace('.', ',')
        q_col = 'D'; vnr_col = 'E'; vos_col = 'F'
        if rep_type == '2':
            ws['D' + str(sec_row - 1)] = 'V1, м3'
            ws['D' + str(sec_row)] = str(round(v1_start, 2)).replace('.', ',')
            ws['D' + str(sec_row + 1)] = str(round(v1_sum, 2)).replace('.', ',')
            ws['E' + str(sec_row - 1)] = 'V2, м3'
            ws['E' + str(sec_row)] = str(round(v2_start, 2)).replace('.', ',')
            ws['E' + str(sec_row + 1)] = str(round(v2_sum, 2)).replace('.', ',')
            for r in range(sec_row - 1, sec_row + 2):
                thin = Side(border_style="thin", color="000000")
                ws.cell(r, 7).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(r, 7).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, 8).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")
            q_col = 'F'; vnr_col = 'G'; vos_col = 'H'
            ws['F' + str(sec_row - 1)] = 'Q, Гкал'
            ws['G' + str(sec_row - 1)] = 'ВНР, час'
            ws['H' + str(sec_row - 1)] = 'ВОС, час'
            
        ws[q_col + str(sec_row)] = str(round(q_start, 2)).replace('.', ',')
        ws[q_col + str(sec_row + 1)] = str(round(q_sum, 2)).replace('.', ',')
        ws[vnr_col + str(sec_row)] = str(round(vnr_start, 2)).replace('.', ',')
        ws[vnr_col + str(sec_row + 1)] = str(round(vnr, 2)).replace('.', ',')
        ws[vos_col + str(sec_row)] = str(round(vos_start, 2)).replace('.', ',')
        ws[vos_col + str(sec_row + 1)] = str(round(vos, 2)).replace('.', ',')

        # Fill head data
        ws['A1'] = str(ws['A1'].value).replace('май', get_month(datetime.now().strftime("%d-%m-%Y")))
        ws['B3'] = date_from
        ws['C3'] = date_to
        ws['B4'] = datetime.now().strftime("%d-%m-%Y")
        ws['B5'] = rep_type
        ws['B6'] = head_data['consumer']
        ws['B7'] = head_data['order']
        ws['B8'] = head_data['adress']
        ws['B11'] = head_data['cold_temp']
        ws['B12'] = head_data['factory_num'].replace('_2', '').replace('_1', '')
        ws['B13'] = head_data['complex_num']
        curr_dir = self.save_dir + '/Output/' + head_data['save_folder']
        if not os.path.exists(curr_dir):
            os.makedirs(curr_dir)
        string_type = '_отопл'
        if rep_type == '2':
            string_type = '_ГВС'
        
        template.save(curr_dir + '/' + head_data['adress'].replace('/', 'к') + string_type + '.xlsx')
        report += curr_dir + '/' + head_data['adress'].replace('/', 'к') + string_type + '.xlsx'+ '\n\n'

        return report


    def __call__(self, date_from = '01-01-2023', date_to = '18-01-2023'): 
        report = '\tВЗЛЕТ\n'   
        for file in self.my_parsing_files:
            rep_type = '1'
            if 'ГВС' in file[0].upper():
                rep_type = '2'
            report += self.build_xls(file, rep_type, date_from, date_to)

        return report
