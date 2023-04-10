from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pyexcel
import os
from datetime import datetime
from HeadData import *


class SPTParser:
    def __init__(self, data_list, curr_dir, save_dir):
        self.report = ''
        self.my_parsing_files = []
        self.my_dir = curr_dir
        self.save_dir = save_dir
        for file in data_list:
            if 'xlsx' not in file:
                if 'xls' in file:
                    pyexcel.save_book_as(file_name=file,
                                dest_file_name=file + 'x')
                    file += 'x'
            if 'xlsx' in file:
                rep_type = '1'
                ws = load_workbook(file).active
                if ws['A1'].value != None:
                    if '_2' in str(ws['A1'].value) or 'ГВС' in str(ws['A1'].value):
                        rep_type = '2'

                if get_head_data(self.my_dir, str(ws['A1'].value), rep_type)['type'] != None:
                    if 'СПТ' in get_head_data(self.my_dir, str(ws['A1'].value), rep_type)['type']:
                        self.my_parsing_files.append([file, load_workbook(file).active])

                    else:
                        continue
                else:
                    continue
            else:
                print('SPT wrong file!')
                continue


    def get_columns(self, row):
        #print([cell.value for cell in row])
        heat_cols = {'Время': 1, 't1(°C)': -1, 't2(°C)': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q(Гкал)': -1, 'Tи(ч)': -1}
        gvs_cols = {'Время': 1, 't1(°C)': -1, 't2(°C)': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q(Гкал)': -1, 'Tи(ч)': -1}
        ind = 0
        for cell in row:
            if cell.value == None:
                ind += 1
                continue
            if 'ТВ1' in cell.value:
                for key in heat_cols.keys():
                    if key in cell.value:
                        heat_cols[key] = ind
                
                if 't1(°C)' in cell.value and gvs_cols['t1(°C)'] == -1:
                    heat_cols['t1(°C)'] = ind
                if 't2(°C)' in cell.value and gvs_cols['t2(°C)'] == -1:
                    heat_cols['t2(°C)'] = ind
                if 'V1' in cell.value and gvs_cols['V1'] == -1:
                    heat_cols['V1'] = ind
                if 'V2' in cell.value and gvs_cols['V2'] == -1:
                    heat_cols['V2'] = ind
                if 'M1' in cell.value and gvs_cols['M1'] == -1:
                    heat_cols['M1'] = ind
                if 'M2' in cell.value and gvs_cols['M2'] == -1:
                    heat_cols['M2'] = ind
                if 'Qг(Гкал)' in cell.value and gvs_cols['Q(Гкал)'] == -1:
                    heat_cols['Q(Гкал)'] = ind
                if 'Tи(ч)' in cell.value and gvs_cols['Tи(ч)'] == -1:
                    heat_cols['Tи(ч)'] = ind

            elif 'ТВ2' in cell.value:
                for key in gvs_cols.keys():
                    if key in cell.value:
                        gvs_cols[key] = ind
                
                if 't3(°C)' in cell.value and gvs_cols['t1(°C)'] == -1:
                    gvs_cols['t1(°C)'] = ind
                if 't4(°C)' in cell.value and gvs_cols['t2(°C)'] == -1:
                    gvs_cols['t2(°C)'] = ind
                if 'V3' in cell.value and gvs_cols['V1'] == -1:
                    gvs_cols['V1'] = ind
                if 'V4' in cell.value and gvs_cols['V2'] == -1:
                    gvs_cols['V2'] = ind
                if 'M3' in cell.value and gvs_cols['M1'] == -1:
                    gvs_cols['M1'] = ind
                if 'M4' in cell.value and gvs_cols['M2'] == -1:
                    gvs_cols['M2'] = ind
                if 'Qг(Гкал)' in cell.value and gvs_cols['Q(Гкал)'] == -1:
                    gvs_cols['Q(Гкал)'] = ind
                if 'Tи(ч)' in cell.value and gvs_cols['Tи(ч)'] == -1:
                    gvs_cols['Tи(ч)'] = ind
            else:
                for key in heat_cols.keys():
                    if key in cell.value:
                        heat_cols[key] = ind

            ind += 1

        return [heat_cols, gvs_cols]


    def get_head(self, filename, report_type):
        filename = filename.split('/')
        filename = filename[len(filename) - 1].replace('.xlsx', '')
        header_ws = load_workbook(self.my_dir + '\Templates\HeadDat.xlsx', data_only=False).active
        head_data = {'factory_num': '', 'complex_num': '', 'consumer': '', 'order': '', 'adress': '', 'cold_temp': '5,0', 'save_folder': ''}
        for i in range(2, 426):
            flag2 = filename.replace('_', ' ')
            if filename.replace('_','').replace('-','').upper() in str(header_ws['C' + str(i)].value).upper() or \
                str(header_ws['C' + str(i)].value).replace(' ', '_') in filename or \
                str(header_ws['E' + str(i)].value).replace('ул.', '') in flag2:
                head_data['factory_num'] = str(header_ws['A' + str(i)].value).replace(str(header_ws['A' + str(i)].value)[len(str(header_ws['A' + str(i)].value))-1], report_type)
                head_data['complex_num'] = header_ws['B' + str(i)].value
                head_data['consumer'] = header_ws['C' + str(i)].value
                head_data['order'] = header_ws['D' + str(i)].value
                head_data['adress'] = header_ws['E' + str(i)].value
                head_data['cold_temp'] = header_ws['G' + str(i)].value
                head_data['save_folder'] = header_ws['J' + str(i)].value

        return head_data
        

    def build_xls(self, file, data_indexes, rep_type, start_row=2):
        date_from = ''
        date_to = ''
        report = ''
        template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx',  read_only=False, data_only=False)  # Template xlsx file  
        file_name = file[0].split('/')[len(file[0].split('/')) - 1].split('.xlsx')[0]
        if file[0].split('/')[len(file[0].split('/')) - 1]:
            template.title = file[0].split('/')[len(file[0].split('/')) - 1]
        ws = template.active

        head_data = {}
        row_index = 1; out_index = 18
        m1_sum = 0; m2_sum = 0; dv_sum = 0; v1_sum = 0; v2_sum = 0; t1_avg = 0; t2_avg = 0; q_sum = 0; vnr = 0; vos = 0; sum_err = ''    
        for row in file[1].iter_rows(min_row=start_row):
            num = lambda t: round(float(str(row[data_indexes[t]].value).replace(',', '.')), 2) if row[data_indexes[t]].value != None else ' - '
            st_row = lambda n: str(n).replace('.', ',')
            if row_index == 1:
                row_index += 1
                continue
            if row[0].value == None:
                break
            curr_date = datetime.date(row[0].value)
            if date_from == '':
                date_from = curr_date
            date_to = curr_date

            ws.insert_rows(out_index)
            for i in range(1, 14):
                thin = Side(border_style="thin", color="000000")
                ws.cell(out_index, i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(out_index, i).alignment = Alignment(horizontal="center", vertical="center")
            ws['A' + str(out_index)] = str(curr_date.strftime("%d-%m-%Y"))
            if data_indexes['t1(°C)'] != -1 and num('t1(°C)') != ' - ':
                t1_avg += num('t1(°C)') 
                ws['B' + str(out_index)] = st_row(num('t1(°C)'))
            if data_indexes['t2(°C)'] != -1 and num('t2(°C)') != ' - ':
                t2_avg += num('t2(°C)')
                ws['C' + str(out_index)] = st_row(num('t2(°C)'))
            if data_indexes['V1'] != -1 and num('V1') != ' - ':
                v1_sum += num('V1')
                ws['D' + str(out_index)] = st_row(num('V1'))
                ws['D' + str(out_index + 1)] = str(round(v1_sum, 2)).replace('.', ',')
                ws['H' + str(out_index)] = st_row(round(num('V1'), 2))
                ws['H' + str(out_index + 1)] = str(abs(round(v1_sum, 2))).replace('.', ',')
            if data_indexes['M1'] != -1 and num('M1') != ' - ': 
                m1_sum += num('M1')
                ws['E' + str(out_index)] = st_row(num('M1'))
                ws['E' + str(out_index + 1)] = str(round(m1_sum, 2)).replace('.', ',')
                ws['I' + str(out_index)] = st_row(num('M1'))
                ws['I' + str(out_index + 1)] = str(abs(round(m1_sum, 2))).replace('.', ',')
            if data_indexes['V2'] != -1 and num('V2') != ' - ':
                v2_sum += num('V2')
                ws['F' + str(out_index)] = st_row(num('V2'))
                ws['F' + str(out_index + 1)] = str(round(v2_sum, 2)).replace('.', ',')
                ws['H' + str(out_index)] = st_row(round(num('V1') - num('V2'), 2))
                ws['H' + str(out_index + 1)] = str(round(v1_sum - v2_sum, 2)).replace('.', ',')
            if data_indexes['M2'] != -1 and num('M2') != ' - ': 
                m2_sum += num('M2')
                ws['G' + str(out_index)] = st_row(num('M2'))
                ws['G' + str(out_index + 1)] = str(round(m2_sum, 2)).replace('.', ',')
                ws['I' + str(out_index)] = st_row(round(num('M1') - num('M2'), 2))
                ws['I' + str(out_index + 1)] = str(round(m1_sum - m2_sum, 2)).replace('.', ',')
            if data_indexes['Q(Гкал)'] != -1 and num('Q(Гкал)') != ' - ':
                q_sum += num('Q(Гкал)')
                ws['J' + str(out_index)] = st_row(num('Q(Гкал)'))
                ws['J' + str(out_index + 1)] = str(round(q_sum, 2)).replace('.', ',')
            if data_indexes['Tи(ч)'] != -1 and num('Tи(ч)') != ' - ': 
                vnr += num('Tи(ч)')
                ws['K' + str(out_index)] = st_row(num('Tи(ч)'))
                ws['K' + str(out_index + 1)] = str(round(vnr, 2)).replace('.', ',')
            if data_indexes['Tи(ч)'] != -1 and num('Tи(ч)') != ' - ': 
                vos += (24.0 - num('Tи(ч)'))
                ws['L' + str(out_index)] = st_row(24.0 - num('Tи(ч)'))
                ws['L' + str(out_index + 1)] = str(round(vos, 2)).replace('.', ',')

            out_index += 1
            row_index += 1

        ws['B' + str(out_index + 1)] = round(t1_avg/(out_index - 18), 2)
        ws['C' + str(out_index + 1)] = round(t2_avg/(out_index - 18), 2)

        # A resoult table 
        sec_row = out_index + 4
        ws['A' + str(sec_row)] = datetime.strftime(date_from, "%d-%m-%Y")
        ws['A' + str(sec_row + 1)] = datetime.strftime(date_to, "%d-%m-%Y")
        ws['B' + str(sec_row)] = '0'
        ws['B' + str(sec_row + 1)] = str(round(m1_sum, 2)).replace('.', ',')
        ws['C' + str(sec_row)] = '0'
        ws['C' + str(sec_row + 1)] = str(round(m2_sum, 2)).replace('.', ',')
        q_col = 'D'; vnr_col = 'E'; vos_col = 'F'
        if rep_type == '2':
            ws['D' + str(sec_row - 1)] = 'V1, м3'
            ws['D' + str(sec_row)] = '0'
            ws['D' + str(sec_row + 1)] = str(round(v1_sum, 2)).replace('.', ',')
            ws['E' + str(sec_row - 1)] = '0'
            ws['E' + str(sec_row)] = '0'
            ws['E' + str(sec_row + 1)] = str(round(v2_sum, 2)).replace('.', ',')
            for r in range(sec_row - 1, sec_row + 2):
                thin = Side(border_style="thin", color="000000")
                ws.cell(r, 7).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(r, 7).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, 8).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")
            q_col = 'F'; vnr_col = 'G'; vos_col = 'H'
            ws['F' + str(sec_row - 1)] = 'Q, Гкал'
            ws['G' + str(sec_row - 1)] = 'ВНР, час'
            ws['H' + str(sec_row - 1)] = 'ВОС, час'
            
        ws[q_col + str(sec_row)] = '0'
        ws[q_col + str(sec_row + 1)] = str(round(q_sum, 2)).replace('.', ',')
        ws[vnr_col + str(sec_row)] = '0'
        ws[vnr_col + str(sec_row + 1)] = str(round(vnr, 2)).replace('.', ',')
        ws[vos_col + str(sec_row)] = '0'
        ws[vos_col + str(sec_row + 1)] = str(round(vos, 2)).replace('.', ',')

        # Fill head data
        if file[1][1][0].value != None:
            if 'Время' in file[1][1][0].value:
                head_data = self.get_head(file[0], rep_type)
            else:
                head_data = get_head_data(self.save_dir, file[1][1][0].value, rep_type)

        ws['A1'] = str(ws['A1'].value).replace('май', get_month(datetime.strftime(date_to, "%d-%m-%Y")))
        ws['B3'] = date_from
        ws['C3'] = date_to
        ws['B4'] = datetime.now().strftime("%d-%m-%Y")
        ws['B5'] = rep_type
        ws['B6'] = head_data['consumer']
        ws['B7'] = head_data['order']
        ws['B8'] = head_data['adress']
        ws['B11'] = head_data['cold_temp']
        ws['B12'] = head_data['factory_num'].replace('_2', '').replace('_1', '')
        ws['B13'] = head_data['complex_num']
        curr_dir = self.save_dir + '/Output/' + head_data['save_folder']
        if not os.path.exists(curr_dir):
            os.makedirs(curr_dir)
        
        name = head_data['consumer'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','') + \
            ' - ' + head_data['adress'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','')
        while name in self.report:
            name += '_2'
        if rep_type == '1':
            template.save(curr_dir + '/' + name + '_отопл' + '.xlsx')
            report += head_data['save_folder'] + '/' + name + '_отопл' + '.xlsx'+ '\n\n'
        else:
            template.save(curr_dir + '/' + name + '_ГВС' + '.xlsx')
            report += head_data['save_folder'] + '/' + name + '_ГВС' + '.xlsx'+ '\n\n'

        return report


    def __call__(self):
        self.report = '\tСПТ:\n' # Window print
        print(len(self.my_parsing_files))
        for file in self.my_parsing_files:
            start_row = 1
            if file[1]['A1'].value != None:
                if 'Время' in file[1]['A1'].value:
                    heat_cols, gvs_cols = self.get_columns(list(file[1].rows)[0])
                    start_row = 1
                else:
                    heat_cols, gvs_cols = self.get_columns(list(file[1].rows)[1])
                    start_row = 2

                if '_1' in str(file[1]['A1'].value):
                    self.report += self.build_xls(file, heat_cols, '1', start_row)
                elif '_2' in str(file[1]['A1'].value):
                    self.report += self.build_xls(file, heat_cols, '2', start_row)
                else: # '_' not in str(file[1]['A1'].value):
                    self.report += self.build_xls(file, heat_cols, '1', start_row)
                    self.report += self.build_xls(file, gvs_cols, '2', start_row)

        return self.report