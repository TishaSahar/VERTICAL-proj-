from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pyexcel
import os
from datetime import datetime
from SPTParser import SPTParser
from HeadData import *

class MKTSParser:
    def __init__(self, data_list, curr_dir, save_dir):
        self.report = ''
        self.my_parsing_files = []
        self.my_dir = curr_dir
        self.save_dir = save_dir
        for file in data_list:
            if 'xlsx' not in file:
                if 'xls' in file:
                    pyexcel.save_book_as(file_name=file,
                                    dest_file_name=file + 'x')
                else:
                    print('MKTS wrong file!')
                    continue
                file += 'x'
            
            rep_type = '1'
            ws = load_workbook(filename=file,  read_only=True).active
            if ws['A1'].value != None:
                if '№' in str(ws['A1'].value) and ',' in str(ws['A1'].value):
                    A1 = str(ws['A1'].value).split('№')[1].split(',')[0]
                else:
                    if '_1' in str(ws['A1'].value) or '_2' in str(ws['A1'].value):
                        if get_head_data(self.my_dir, str(ws['A1'].value), rep_type)['type'] != None:
                            if 'МКТС' in get_head_data(self.my_dir, str(ws['A1'].value).replace('-1', '').replace('-2', ''), rep_type)['type']:
                                self.my_parsing_files.append([file, load_workbook(filename=file,  read_only=True).active])
                            else:
                                continue
                    continue
                if '-2' in str(A1) or 'ГВС' in file:
                    rep_type = '2'

                #print(get_head_data(self.my_dir, str(A1), rep_type)['type'])
                if get_head_data(self.my_dir, str(A1), rep_type)['type'] != None:
                    if ('МКТС' in get_head_data(self.my_dir, str(A1), rep_type)['type']) or ('МКТС' in str(ws['A1'].value)):
                        self.my_parsing_files.append([file, load_workbook(filename=file,  read_only=True).active])
                    else:
                        continue
                else:
                    print('Can not to expect type in: ',  str(file))
                    continue
            else:
                continue
                #print('Can not to expect factory num in: ',  str(file))


    def data_index(self, row):
        nums_of_data = {'Дат': -1, 't1,°С': -1, 't2,°С': -1,'V1': -1,'M1': -1,'V2': -1,'M2': -1, 'Q': -1, 'Tраб': -1, 'Tотк': -1, 'Отка': -1}
        for key in nums_of_data.keys():
            ind = 0
            for cell in row:
                if key in cell.value:
                    nums_of_data[key] = ind
                    break
                ind += 1

        return nums_of_data


    def get_err(self, value, sum_err):
        ret = ''
        errs = {'Не': '3', 'DG': '4', 'ДG': '4', 'Dt': '3', 'Эл': '1'}
        for err in errs.keys():
            if err in value and err not in ret:
                ret = errs[err] if ret == '' else ret + ', ' + errs[err]
                if sum_err == '':
                    sum_err = ret
                elif errs[err] not in sum_err:
                    sum_err += ', ' + errs[err]
                    
        return ret, sum_err


    def __call__(self):
        self.report = '\tМКТС\n' # Window print
        print(len(self.my_parsing_files))
        for file in self.my_parsing_files:
            if file[1]['A1'].value != None:
                if 'Время' in str(file[1]['A2'].value):
                    spt_parser = SPTParser([], self.my_dir, self.save_dir)
                    columns, gvs_cols = spt_parser.get_columns(list(file[1].rows)[1])

                    if '_1' in str(file[1]['A1'].value):
                        self.report += spt_parser.build_xls(file, columns, '1', 2)
                    elif '_2' in str(file[1]['A1'].value):
                        self.report += spt_parser.build_xls(file, columns, '2', 2)
                    else: # '_' not in str(file[1]['A1'].value):
                        self.report += spt_parser.build_xls(file, columns, '1', 2)
                        self.report += spt_parser.build_xls(file, columns, '2', 2)

                    continue

            date_from = ''
            date_to = ''
            template = load_workbook(self.my_dir + '\Templates\VEC_Template.xlsx',  read_only=False, data_only=False)  # Template xlsx file
            if file[0].split('/')[len(file[0].split('/')) - 1]:
                template.title = file[0].split('/')[len(file[0].split('/')) - 1]
            ws = template.active

            order_type = '1'
            if 'ГВС' in file[0].upper():
                order_type = '2'

            head_data = {}
            data_index = {}
            row_index = 1
            out_index = 18

            dm_sum = 0
            m1_sum = 0
            m2_sum = 0 
            dv_sum = 0
            v1_sum = 0
            v2_sum = 0 
            t1_avg = 0
            t2_avg = 0
            q_sum = 0
            vnr = 0
            vos = 0        
            sum_err = ''    
            for row in file[1].iter_rows():
                num = lambda t: round(float(row[data_index[t]].value), 2)
                st_row = lambda n: str(n).replace('.', ',')
                if row_index == 1: 
                    head_data = get_head_data(self.my_dir, row[0].value.split('№')[1].split(',')[0], order_type)
                    row_index += 1
                    continue
                if row_index == 2:
                    data_index = self.data_index(row)
                    row_index += 1
                    continue
                if 'Итого' in row[0].value:
                    break
    
                tmp_date = row[0].value.split(',')
                curr_date = datetime.strptime(tmp_date[0] + '-' + tmp_date[1] + '-' + '20' + tmp_date[2], "%d-%m-%Y").date()
                if date_from == '':
                    date_from = curr_date
                date_to = curr_date
                ws.insert_rows(out_index)
                for i in range(1, 14):
                    thin = Side(border_style="thin", color="000000")
                    ws.cell(out_index, i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(out_index, i).alignment = Alignment(horizontal="center", vertical="center")
                #[print(row[i].value, ' ') for i in range(0, 15)]
                ws['A' + str(out_index)] = str(curr_date.strftime("%d-%m-%Y"))
                if data_index['t1,°С'] == -1 or '-' in str(row[data_index['t1,°С']].value):
                    ws['B' + str(out_index)] = ' - '
                else:
                    ws['B' + str(out_index)] = st_row(num('t1,°С'))
                    t1_avg += num('t1,°С')
                if data_index['t2,°С'] == -1 or '-' in str(row[data_index['t2,°С']].value):
                    ws['C' + str(out_index)] = ' - '
                else:
                    ws['C' + str(out_index)] = st_row(num('t2,°С'))
                    t2_avg += num('t2,°С')
                if data_index['V1'] == -1 or '-' in str(row[data_index['V1']].value):
                    ws['D' + str(out_index)] = ' - '
                else:
                    ws['D' + str(out_index)] = st_row(num('V1'))
                    v1_sum += num('V1')
                    ws['D' + str(out_index + 1)] = st_row(round(v1_sum))
                if data_index['M1'] == -1 or '-' in str(row[data_index['M1']].value):
                    ws['E' + str(out_index)] = ' - '
                else:
                    ws['E' + str(out_index)] = st_row(num('M1'))
                    m1_sum += num('M1')
                    dm_sum = m1_sum
                    ws['E' + str(out_index + 1)] = st_row(round(m1_sum, 2))
                if data_index['V2'] == -1 or '-' in str(row[data_index['V2']].value):
                    ws['F' + str(out_index)] = ' - '
                    ws['H' + str(out_index)] = ws['D' + str(out_index)].value
                else:
                    ws['F' + str(out_index)] = st_row(num('V2'))
                    v2_sum += num('V2')
                    dv_sum += num('V1') - num('V2')
                    ws['F' + str(out_index + 1)] = st_row(round(v2_sum, 2))
                    ws['H' + str(out_index)] = st_row(round(num('V1') - num('V2'), 2))
                    ws['H' + str(out_index + 1)] = st_row(round(dv_sum, 2))
                if data_index['M2'] == -1 or '-' in str(row[data_index['M2']].value):
                    ws['G' + str(out_index)] = ' - '
                    ws['I' + str(out_index)] = ws['E' + str(out_index)].value
                    ws['I' + str(out_index + 1)] = round(dm_sum, 2)
                else:
                    ws['G' + str(out_index)] = st_row(num('M2'))
                    m2_sum += num('M2')
                    dm_sum = round(m1_sum - m2_sum, 2)
                    ws['G' + str(out_index + 1)] = st_row(round(m2_sum, 2))
                    ws['I' + str(out_index)] = st_row(round(num('M1') - num('M2'), 2))
                    ws['I' + str(out_index + 1)] = st_row(dm_sum)
                if data_index['Q'] == -1  or '-' in str(row[data_index['Q']].value):
                    ws['J' + str(out_index)] = ' - '
                else:
                    ws['J' + str(out_index)] = st_row(num('Q'))
                    q_sum += num('Q')
                    ws['J' + str(out_index + 1)] = st_row(round(q_sum, 3))
                
                if data_index['Tраб'] == -1 or '-' in str(row[data_index['Tраб']].value):
                    ws['K' + str(out_index)] = ' - '
                else:
                    ws['k' + str(out_index)] = st_row(num('Tраб'))
                    vnr += num('Tраб')
                    ws['K' + str(out_index + 1)] = st_row(round(vnr, 2))
                if data_index['Tотк'] == -1 or '-' in str(row[data_index['Tотк']].value):
                    ws['L' + str(out_index)] = ' - '
                else:
                    ws['L' + str(out_index)] = st_row(num('Tотк'))
                    vos += num('Tотк')
                    ws['L' + str(out_index + 1)] = st_row(round(vos, 2))
                if data_index['Отка'] == -1 or type(row[data_index['Отка']].value) != str:
                    ws['M' + str(out_index)] = ' '
                else:
                    err, sum_err = self.get_err(str(row[data_index['Отка']].value), sum_err)
                    ws['M' + str(out_index)] = err
                    #if err not in sum_err:
                    #    sum_err = err if sum_err == '' else sum_err + ', ' + err
                    ws['M' + str(out_index + 1)] = sum_err

                out_index += 1
                row_index += 1

            if t1_avg != 0:
                ws['B' + str(out_index + 1)] = str(round(t1_avg / (out_index-18), 2)).replace('.', ',')
            if t2_avg != 0:
                ws['C' + str(out_index + 1)] = str(round(t2_avg / (out_index-18), 2)).replace('.', ',')
            
            # A resoult table 
            q_col = 'D'
            vnr_col = 'E'
            vos_col = 'F'
            
            sec_row = out_index + 4
            row_index += 4
            final_m1 = '-'; final_m2 ='-'; final_v1 = '-'; final_v2 = '-'; final_q = '-'; final_vnr = '-'; final_vos = '-'
            if data_index['M1'] != -1:
                if '-' not in str(file[1].cell(row=row_index, column=data_index['M1'] + 1).value):
                    final_m1 = float(str(file[1].cell(row=row_index, column=data_index['M1'] + 1).value).replace(',', '.').replace(' ', ''))
            if data_index['M2'] != -1:
                if '-' not in str(file[1].cell(row=row_index, column=data_index['M2'] + 1).value):
                    final_m2 = float(str(file[1].cell(row=row_index, column=data_index['M2']+1).value).replace(',', '.').replace(' ', ''))
            if data_index['V1'] != -1:
                if '-' not in str(file[1].cell(row=row_index, column=data_index['V1'] + 1).value):
                    final_v1 = float(str(file[1].cell(row=row_index, column=data_index['V1'] + 1).value).replace(',', '.').replace(' ', ''))
            else:
                final_v1 = '-'
            if data_index['V2'] != -1: 
                if '-' not in str(file[1].cell(row=row_index, column=data_index['V2'] + 1).value):
                    final_v2 = float(str(file[1].cell(row=row_index, column=data_index['V2'] + 1).value).replace(',', '.').replace(' ', ''))

            if data_index['Q'] != -1: 
                if '-' not in str(file[1].cell(row=row_index, column=data_index['Q']+1).value):
                    final_q = float(str(file[1].cell(row=row_index, column=data_index['Q'] + 1).value).replace(',', '.').replace(' ', ''))

            if data_index['Tраб'] != -1:
                if '-' not in str(file[1].cell(row=row_index, column=data_index['Tраб']+1).value):
                    final_vnr = float(str(file[1].cell(row=row_index, column=data_index['Tраб']+1).value).replace(',', '.').replace(' ', ''))
            if data_index['Tотк'] != -1: 
                if '-' not in str(file[1].cell(row=row_index, column=data_index['Tотк']+1).value):  
                    final_vos = float(str(file[1].cell(row=row_index, column=data_index['Tотк']+1).value).replace(',', '.').replace(' ', ''))
            #if data_index['Отка'] != -1 :
            #    if '-' not in str(file[1].cell(row=row_index, column=data_index['Отка']+1).value):
            #        sum_err = str(file[1].cell(row=row_index, column=data_index['Отка']+1).value)
            ws['A' + str(sec_row)] = date_from
            ws['A' + str(sec_row + 1)] = date_to
            if final_m1 != '-':
                ws['B' + str(sec_row)] = final_m1
                ws['B' + str(sec_row + 1)] = str(round(m1_sum + final_m1, 2)).replace('.', ',')
            else:
                ws['B' + str(sec_row)] = 0
                ws['B' + str(sec_row + 1)] = str(round(m1_sum, 2)).replace('.', ',')
            if final_m2 != '-':
                ws['C' + str(sec_row)] = final_m2
                ws['C' + str(sec_row + 1)] = str(round(m2_sum + final_m2, 2)).replace('.', ',')
            else:
                ws['C' + str(sec_row)] = 0
                ws['C' + str(sec_row + 1)] = str(round(m2_sum, 2)).replace('.', ',')

            if str(order_type) == '2':
                ws['D' + str(sec_row - 1)] = 'V1, м3'
                if final_v1 != '-':
                    ws['D' + str(sec_row)] = final_v1
                    ws['D' + str(sec_row + 1)] = str(round(v1_sum + final_v1, 2)).replace('.', ',')
                else:
                    ws['D' + str(sec_row)] = 0
                    ws['D' + str(sec_row + 1)] = str(round(v1_sum, 2)).replace('.', ',')

                ws['E' + str(sec_row - 1)] = 'V2, м3'
                if final_v2 != '-':
                    ws['E' + str(sec_row)] = final_v2
                    ws['E' + str(sec_row + 1)] = str(round(v2_sum + final_v2, 2)).replace('.', ',')
                else:
                    ws['E' + str(sec_row)] = 0
                    ws['E' + str(sec_row + 1)] = str(round(v2_sum, 2)).replace('.', ',')

                ws['F' + str(sec_row - 1)] = 'Q, Гкал'
                ws['G' + str(sec_row - 1)] = 'ВНР, час'
                ws['H' + str(sec_row - 1)] = 'ВОС, час'
                q_col = 'F'
                vnr_col = 'G'
                vos_col = 'H'
                for r in range(sec_row - 1, sec_row + 2):
                    thin = Side(border_style="thin", color="000000")
                    ws.cell(r, 7).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(r, 7).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(r, 8).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")
            
            if final_q != '-':
                ws[q_col + str(sec_row)] = final_q
                ws[q_col + str(sec_row + 1)] = str(round(q_sum + final_q, 2)).replace('.', ',')
            else:
                ws[q_col + str(sec_row)] = 0
                ws[q_col + str(sec_row + 1)] = str(round(q_sum, 2)).replace('.', ',')

            if final_vnr != '-':
                ws[vnr_col + str(sec_row)] = final_vnr
                ws[vnr_col + str(sec_row + 1)] = str(round(vnr + final_vnr, 2)).replace('.', ',')
            else:
                ws[vnr_col + str(sec_row)] = 0
                ws[vnr_col + str(sec_row + 1)] = str(round(vnr, 2)).replace('.', ',')
            
            if final_vos != '-':
                ws[vos_col + str(sec_row)] = final_vos
                ws[vos_col + str(sec_row + 1)] = str(round(vos + final_vos, 2)).replace('.', ',')
            else:
                ws[vos_col + str(sec_row)] = 0
                ws[vos_col + str(sec_row + 1)] = str(round(vos, 2)).replace('.', ',')

            # Fill head data
            ws['A1'] = str(ws['A1'].value).replace('май', get_month(str(date_to)[8:10] + '-' + str(date_to)[5:7] + '-' + str(date_to)[0:4]))
            ws['B3'] = datetime.strftime(date_from, "%d-%m-%Y")
            ws['C3'] = datetime.strftime(date_to, "%d-%m-%Y")
            ws['B4'] = datetime.now().strftime("%d-%m-%Y")
            ws['B5'] = order_type
            ws['B6'] = head_data['consumer']
            ws['B7'] = head_data['order']
            ws['B8'] = head_data['adress']
            ws['B11'] = head_data['cold_temp']
            ws['B12'] = head_data['factory_num']
            ws['B13'] = head_data['complex_num']

            curr_dir = self.save_dir + '/Output/' + head_data['save_folder']
            if not os.path.exists(curr_dir):
                os.makedirs(curr_dir)

            str_type = '_отопл'
            if 'ГВС' in file[0].upper():
                str_type = '_ГВС'

            name = head_data['consumer'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','') + \
            ' - ' + head_data['adress'].replace(',', '').replace('/', 'к').replace('"', '').replace('<','').replace('>','').replace('?','').replace('*','').replace('|','')
            while name in self.report:
                name += '_2'
            template.save(curr_dir + '/' + name + str_type + '.xlsx')
            self.report += head_data['save_folder'] + '/' + name + str_type + '.xlsx'+ '\n\n'

        return self.report